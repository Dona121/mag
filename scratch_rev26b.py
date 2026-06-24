# -*- coding: utf-8 -*-
import sys; sys.stdout.reconfigure(encoding="utf-8")
import openpyxl, re
from collections import defaultdict
from contenido.models import DependenciaModelo, Subindicador, Criterio
P = r"C:\Users\Donal\Documents\MEGA\Estudio\Django\modelo_alta_gerencia\migracion\estructura_modelo_version_1_2026.xlsx"
wb = openpyxl.load_workbook(P, data_only=True)
CANON={"Ciclo de Proyectos":"Ciclos de Gerencia","Mecanismos de Financiación":"Otros Mecanismos de Financiación"}
def cp(x): return CANON.get(x,x)
def clean(v): return re.sub(r"\s+"," ",str(v).replace("\xa0"," ")).strip() if v is not None else None
def f(v):
    try: return float(v)
    except (TypeError,ValueError): return None

dep_model={dm.dependencia.nombre:dm.modelo_id for dm in DependenciaModelo.objects.filter(modelo__version=1).select_related("dependencia")}
db_names=set()  # nombres de catálogo en BD v1 (pilar/ind/sub)
for r in Subindicador.objects.filter(indicador__pilar__modelo_evaluacion__version=1).values(
        "indicador__pilar__nombre__nombre","indicador__nombre__nombre","nombre__nombre"):
    db_names.update([r["indicador__pilar__nombre__nombre"],r["indicador__nombre__nombre"],r["nombre__nombre"]])
db_crit=set(clean(x) for x in Criterio.objects.filter(subindicador__indicador__pilar__modelo_evaluacion__version=1).values_list("nombre",flat=True))

# nombres del 2026
xl_pil=set(); xl_ind=set(); xl_sub=set(); xl_crit=set()
for ws in wb.worksheets:
    if ws.title.strip().lower()=="categorias": continue
    for r in range(2, ws.max_row+1):
        pn=clean(ws.cell(r,1).value); inn=clean(ws.cell(r,3).value); sn=clean(ws.cell(r,5).value); cn=clean(ws.cell(r,7).value)
        if pn: xl_pil.add(cp(pn))
        if inn: xl_ind.add(inn)
        if sn: xl_sub.add(sn)
        if cn: xl_crit.add(cn)

print("=== [B] Nombres en 2026 que NO existen en el catálogo v1 de la BD ===")
for label,s in [("PILARES",xl_pil),("INDICADORES",xl_ind),("SUBINDICADORES",xl_sub)]:
    nuevos=[x for x in s if x not in db_names]
    print(f"\n  {label}: {len(nuevos)} no coinciden con v1")
    for x in sorted(nuevos): print(f"     ! {x!r}")
nuevos_c=[x for x in xl_crit if x not in db_crit]
print(f"\n  CRITERIOS: {len(nuevos_c)} no coinciden con v1 (muestra 15)")
for x in sorted(nuevos_c)[:15]: print(f"     ! {x!r}")

print("\n=== Dump 'Desarrollo Económico' (col1/3/5) — dio pilar None ===")
ws=wb["Desarrollo Económico"]
for r in range(1, min(ws.max_row,12)+1):
    print(f"   fila{r}: pilar={clean(ws.cell(r,1).value)!r} ind={clean(ws.cell(r,3).value)!r} sub={clean(ws.cell(r,5).value)!r}")
