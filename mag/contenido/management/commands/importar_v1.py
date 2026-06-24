# -*- coding: utf-8 -*-
"""
Importa la versión 1 (histórica, llevada en Excel) del Modelo de Alta Gerencia.

Estructura del archivo .xlsx:
  - Hoja 'categorias' = mapa Dependencia -> Categoría (Categoría 1/2/3).
  - Las demás hojas: 1 hoja = 1 Dependencia (el título de la hoja es su nombre).
  - Fila 1 = encabezados.  Filas 2..n = jerarquía aplanada
    Pilar > Indicador > Subindicador > Criterio (con relleno hacia abajo: la celda
    vacía hereda el valor de la fila superior).
  - Columnas 1..8  = catálogo del modelo (nombre/peso de cada nivel + criterio/rango).
  - Columnas 9..29 = resultados pivoteados por periodo (vigencia 2025).
  - Columna 30     = tipo_calculo del subindicador.

Cada dependencia evaluó su propio subconjunto de pilares/indicadores/subindicadores.
Se agrupan las dependencias por ESTRUCTURA idéntica (el árbol, tras normalizar
nombres) y se crea un ModeloEvaluacion v1 por estructura.  El catálogo (pesos y
criterios) de cada modelo se toma de la PRIMERA dependencia del grupo; las
divergencias de pesos/criterios/tipo_calculo dentro del grupo se reportan.  Los
resultados (puntaje/ponderación) se guardan directo de cada hoja, así que el dato
histórico de cada dependencia queda exacto aunque el catálogo use la representativa.

Todo lo numérico viene en fracción (0..1) y se guarda como porcentaje (x100),
para igualar la convención ya presente en la BD (peso 30.00, puntaje 86.67, etc.).

Por defecto corre en modo SIMULACIÓN (no escribe nada). Use --commit para escribir.
"""
import re
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from contenido.models import (
    Categoria,
    Criterio,
    Dependencia,
    DependenciaModelo,
    Evaluacion,
    EvaluacionResultado,
    EvaluacionResultadoDetalle,
    Indicador,
    IndicadorCategoria,
    ModeloEvaluacion,
    Periodo,
    Pilar,
    PilarCategoria,
    Subindicador,
    SubindicadorCategoria,
)

# --- Configuración de la migración -------------------------------------------

VERSION = 1
VIGENCIA = 2025
HOJA_CATEGORIAS = "categorias"

# Nombres canónicos (variantes que son el mismo concepto).
CANON_PILAR = {
    "Ciclo de Proyectos": "Ciclos de Gerencia",
    "Mecanismos de Financiación": "Otros Mecanismos de Financiación",
}
# Hojas cuyo título no es el nombre definitivo de la dependencia.
CANON_DEPENDENCIA = {
    "Oficina TI": "Oficina TIC",
}

# Periodos (orden, nombre, [(mes, col_puntaje, col_ponderado)], col_pond_periodo, col_obs)
# El periodo Enero - Junio solo trae junio, así que su ponderación de periodo ES la
# ponderación de junio.  Columnas en base 1 (como en Excel).
PERIODOS = [
    (1, "Enero - Junio",          [(6, 9, 10)],                    10, 11),
    (2, "Julio - Agosto",         [(7, 12, 13), (8, 14, 15)],      16, 17),
    (3, "Septiembre - Octubre",   [(9, 18, 19), (10, 20, 21)],     22, 23),
    (4, "Noviembre - Diciembre",  [(11, 24, 25), (12, 26, 27)],    28, 29),
]

# 5 decimales para igualar DecimalField(max_digits=10, decimal_places=5) del modelo.
Q5 = Decimal("0.00001")

# --- Limpieza de texto -------------------------------------------------------


def limpiar_nombre(valor):
    """Normaliza un nombre de catálogo: \\xa0->espacio, colapsa espacios, strip y typos."""
    if valor is None:
        return None
    texto = str(valor).replace("\xa0", " ")
    texto = re.sub(r"\s+", " ", texto).strip()
    texto = texto.replace("cumplimeinto", "cumplimiento")
    texto = re.sub(r"\bCumplimento\b", "Cumplimiento", texto)
    return texto


def canon_pilar(nombre):
    return CANON_PILAR.get(nombre, nombre)


def limpiar_obs(valor):
    """Para observaciones: solo normaliza \\xa0 y recorta extremos (conserva saltos)."""
    if valor is None:
        return ""
    return str(valor).replace("\xa0", " ").strip()


def norm_tipo_calculo(valor):
    """Normaliza tipo_calculo a las opciones del modelo: 'mensual' | 'directo'."""
    texto = limpiar_nombre(valor)
    if not texto:
        return None
    bajo = texto.lower()
    if bajo.startswith("directo"):
        return "directo"
    if bajo.startswith("mensual"):
        return "mensual"
    return bajo


def a_porcentaje(valor):
    """Fracción 0..1 -> porcentaje 0..100 con 2 decimales."""
    if valor is None:
        return None
    return (Decimal(str(valor)) * 100).quantize(Q5, rounding=ROUND_HALF_UP)


def a_rango(valor):
    if valor is None:
        return ""
    if isinstance(valor, str):
        return limpiar_nombre(valor)
    return "%g" % valor  # número (1, 0.8, 0.05): texto sin notación científica


# --- Parseo del Excel --------------------------------------------------------


def parsear_categorias(ws):
    """Hoja 'categorias' -> {dependencia_canonica: nombre_categoria}."""
    mapa = {}
    for r in range(2, ws.max_row + 1):
        dep = limpiar_nombre(ws.cell(r, 1).value)
        cat = limpiar_nombre(ws.cell(r, 2).value)
        if not dep:
            continue
        dep = CANON_DEPENDENCIA.get(dep, dep)
        if cat:
            mapa[dep] = cat
    return mapa


def parsear_hoja(ws):
    """Devuelve (dependencia_canonica, [pilares]) con jerarquía, pesos y resultados."""
    pilares = []
    pilar_act = ind_act = sub_act = None
    orden_pilar = orden_ind = orden_sub = orden_crit = 0

    for r in range(2, ws.max_row + 1):
        def cel(c):
            return ws.cell(r, c).value

        pilar_nom = canon_pilar(limpiar_nombre(cel(1)))
        ind_nom = limpiar_nombre(cel(3))
        sub_nom = limpiar_nombre(cel(5))
        crit_nom = limpiar_nombre(cel(7))

        if pilar_nom:
            orden_pilar += 1
            pilar_act = {
                "orden": orden_pilar,
                "nombre": pilar_nom,
                "peso": a_porcentaje(cel(2)),
                "indicadores": [],
            }
            pilares.append(pilar_act)
            orden_ind = 0

        if ind_nom:
            orden_ind += 1
            ind_act = {
                "orden": orden_ind,
                "nombre": ind_nom,
                "peso": a_porcentaje(cel(4)),
                "subindicadores": [],
            }
            pilar_act["indicadores"].append(ind_act)
            orden_sub = 0

        if sub_nom:
            orden_sub += 1
            sub_act = {
                "orden": orden_sub,
                "nombre": sub_nom,
                "peso": a_porcentaje(cel(6)),
                "tipo_calculo": norm_tipo_calculo(cel(30)),
                "criterios": [],
                "resultados": _leer_resultados(ws, r),
            }
            ind_act["subindicadores"].append(sub_act)
            orden_crit = 0

        if crit_nom:
            orden_crit += 1
            sub_act["criterios"].append(
                {"orden": orden_crit, "nombre": crit_nom, "rango": a_rango(cel(8))}
            )

    dep = ws.title.strip()
    return CANON_DEPENDENCIA.get(dep, dep), pilares


def _leer_resultados(ws, fila):
    """Lee, para la fila ancla de un subindicador, sus resultados por periodo."""
    resultados = []
    for orden_p, nombre_p, meses, col_pond, col_obs in PERIODOS:
        detalles = []
        for mes, col_pun, col_pon in meses:
            pun = ws.cell(fila, col_pun).value
            pon = ws.cell(fila, col_pon).value
            if pun is None and pon is None:
                continue
            detalles.append(
                {
                    "mes": mes,
                    "puntaje": a_porcentaje(pun if pun is not None else 0),
                    "ponderacion": a_porcentaje(pon if pon is not None else 0),
                }
            )
        pond_periodo = ws.cell(fila, col_pond).value
        if not detalles and pond_periodo is None:
            continue  # sin datos en este periodo
        if detalles:  # puntaje del periodo = promedio de los meses con dato
            prom = sum((d["puntaje"] for d in detalles), Decimal("0")) / len(detalles)
            puntaje = prom.quantize(Q5, rounding=ROUND_HALF_UP)
        else:
            puntaje = Decimal("0.00000")
        resultados.append(
            {
                "orden": orden_p,
                "periodo": nombre_p,
                "puntaje": puntaje,
                "ponderacion": a_porcentaje(pond_periodo) or Decimal("0.00000"),
                "observaciones": limpiar_obs(ws.cell(fila, col_obs).value),
                "detalles": detalles,
            }
        )
    return resultados


def firma_estructura(pilares):
    """Tupla (pilar, indicador, subindicador) que identifica la ESTRUCTURA."""
    return tuple(
        (p["nombre"], i["nombre"], s["nombre"])
        for p in pilares
        for i in p["indicadores"]
        for s in i["subindicadores"]
    )


def _indice_catalogo(pilares):
    """{(pilar,indicador,subindicador): {peso,tipo_calculo,criterios,pilar_peso,ind_peso}}"""
    out = {}
    for p in pilares:
        for i in p["indicadores"]:
            for s in i["subindicadores"]:
                out[(p["nombre"], i["nombre"], s["nombre"])] = {
                    "pilar_peso": p["peso"],
                    "ind_peso": i["peso"],
                    "sub_peso": s["peso"],
                    "tipo_calculo": s["tipo_calculo"],
                    "criterios": [(c["nombre"], c["rango"]) for c in s["criterios"]],
                }
    return out


# --- Comando -----------------------------------------------------------------


class Command(BaseCommand):
    help = "Importa la versión 1 (Excel) del modelo. Por defecto simula; use --commit para escribir."

    def add_arguments(self, parser):
        parser.add_argument("ruta", help="Ruta al archivo .xlsx")
        parser.add_argument(
            "--commit",
            action="store_true",
            help="Escribe en la base de datos (sin esta bandera solo simula).",
        )

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover
            raise CommandError("Falta openpyxl: instálelo con 'uv sync'.") from exc

        ruta, commit = opts["ruta"], opts["commit"]
        try:
            wb = openpyxl.load_workbook(ruta, data_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f"No se encontró el archivo: {ruta}") from exc

        modo = "ESCRITURA (--commit)" if commit else "SIMULACIÓN (dry-run)"
        self.stdout.write(self.style.MIGRATE_HEADING(f"\nImportador MAG v{VERSION} — modo {modo}"))
        self.stdout.write(f"Archivo: {ruta}")

        # 1) Categorías (dependencia -> categoría)
        if HOJA_CATEGORIAS not in [s.lower() for s in wb.sheetnames]:
            raise CommandError(f"No se encontró la hoja '{HOJA_CATEGORIAS}'.")
        ws_cat = next(ws for ws in wb.worksheets if ws.title.strip().lower() == HOJA_CATEGORIAS)
        mapa_cat = parsear_categorias(ws_cat)

        # 2) Parsear dependencias y agrupar por estructura
        hojas = [
            parsear_hoja(ws)
            for ws in wb.worksheets
            if ws.title.strip().lower() != HOJA_CATEGORIAS
        ]
        grupos = defaultdict(list)  # firma -> [(dep, pilares)]
        for dep, pilares in hojas:
            grupos[firma_estructura(pilares)].append((dep, pilares))
        # nombre de modelo por estructura
        self.modelos = []
        for idx, (firma, miembros) in enumerate(grupos.items(), 1):
            rep_dep = miembros[0][0]
            nombre = f"MAG v1 - {rep_dep}" + (" (y similares)" if len(miembros) > 1 else "")
            self.modelos.append(
                {
                    "nombre": nombre,
                    "indice": idx,
                    "miembros": miembros,
                    "representativa": miembros[0],  # (dep, pilares)
                }
            )

        self._reportar(mapa_cat)

        if not commit:
            self.stdout.write(
                self.style.WARNING(
                    "\n>>> SIMULACIÓN: no se escribió nada. Reejecute con --commit para aplicar."
                )
            )
            return

        with transaction.atomic():
            self._escribir(mapa_cat)
        self.stdout.write(self.style.SUCCESS("\n>>> Importación aplicada."))

    # -- reporte --------------------------------------------------------------

    def _reportar(self, mapa_cat):
        self.stdout.write(self.style.MIGRATE_LABEL("\nCategorías (hoja 'categorias'):"))
        por_cat = defaultdict(list)
        for dep, cat in mapa_cat.items():
            por_cat[cat].append(dep)
        for cat in sorted(por_cat):
            self.stdout.write(f"  {cat}: {', '.join(sorted(por_cat[cat]))}")

        self.stdout.write(
            self.style.MIGRATE_LABEL(
                f"\nEstructuras detectadas: {len(self.modelos)} (un ModeloEvaluacion v1 c/u)"
            )
        )
        tot_eval = tot_er = tot_det = 0
        for m in self.modelos:
            deps = [d for d, _ in m["miembros"]]
            rep_dep, rep_pilares = m["representativa"]
            n_pil = len(rep_pilares)
            n_ind = sum(len(p["indicadores"]) for p in rep_pilares)
            n_sub = sum(len(i["subindicadores"]) for p in rep_pilares for i in p["indicadores"])
            n_cri = sum(
                len(s["criterios"])
                for p in rep_pilares
                for i in p["indicadores"]
                for s in i["subindicadores"]
            )
            self.stdout.write(self.style.HTTP_INFO(f"\n# {m['nombre']}"))
            self.stdout.write(f"    Dependencias: {', '.join(deps)}")
            self.stdout.write(f"    Representativa (catálogo): {rep_dep}")
            self.stdout.write(
                f"    Catálogo: {n_pil} pilares, {n_ind} indicadores, "
                f"{n_sub} subindicadores, {n_cri} criterios"
            )
            self._reportar_divergencias(m)
            # resultados por dependencia del grupo
            for dep, pilares in m["miembros"]:
                er = det = 0
                periodos = set()
                for p in pilares:
                    for i in p["indicadores"]:
                        for s in i["subindicadores"]:
                            for res in s["resultados"]:
                                er += 1
                                det += len(res["detalles"])
                                periodos.add(res["periodo"])
                cat = mapa_cat.get(dep, "(sin categoría!)")
                self.stdout.write(
                    f"      - {dep} [{cat}]: {len(periodos)} periodos -> "
                    f"{len(periodos)} evaluaciones, {er} resultados, {det} detalles"
                )
                tot_eval += len(periodos)
                tot_er += er
                tot_det += det

        self.stdout.write(self.style.HTTP_INFO(
            f"\nTOTALES a crear: {len(self.modelos)} modelos, "
            f"{len(mapa_cat)} dependencias (mapeadas), {tot_eval} evaluaciones, "
            f"{tot_er} EvaluacionResultado, {tot_det} EvaluacionResultadoDetalle."
        ))
        self._reportar_existentes(mapa_cat)

    def _reportar_divergencias(self, modelo):
        rep_dep, rep_pilares = modelo["representativa"]
        base = _indice_catalogo(rep_pilares)
        avisos = []
        for dep, pilares in modelo["miembros"][1:]:
            idx = _indice_catalogo(pilares)
            for key, val in idx.items():
                b = base.get(key)
                if not b:
                    continue
                sub = key[2][:30]
                if val["pilar_peso"] != b["pilar_peso"]:
                    avisos.append(f"{dep}: peso pilar '{key[0]}' {val['pilar_peso']} vs {b['pilar_peso']}")
                if val["ind_peso"] != b["ind_peso"]:
                    avisos.append(f"{dep}: peso indicador '{key[1]}' {val['ind_peso']} vs {b['ind_peso']}")
                if val["sub_peso"] != b["sub_peso"]:
                    avisos.append(f"{dep}: peso sub '{sub}' {val['sub_peso']} vs {b['sub_peso']}")
                if val["tipo_calculo"] != b["tipo_calculo"]:
                    avisos.append(f"{dep}: tipo_calculo '{sub}' {val['tipo_calculo']} vs {b['tipo_calculo']}")
                if val["criterios"] != b["criterios"]:
                    avisos.append(f"{dep}: criterios distintos en '{sub}'")
        # de-dup conservando orden
        vistos = set()
        unicos = [a for a in avisos if not (a in vistos or vistos.add(a))]
        if unicos:
            self.stdout.write(self.style.WARNING(
                f"    Divergencias vs representativa (se usa la de {rep_dep}):"
            ))
            for a in unicos:
                self.stdout.write(f"        ! {a}")

    def _reportar_existentes(self, mapa_cat):
        self.stdout.write(self.style.MIGRATE_LABEL("\nVerificación en la BD actual:"))
        for m in self.modelos:
            ya = ModeloEvaluacion.objects.filter(nombre=m["nombre"], version=VERSION).first()
            estado = f"YA EXISTE (pk={ya.pk}, se reutiliza)" if ya else "se creará"
            self.stdout.write(f"  - {m['nombre']}: {estado}")
        cats = sorted(set(mapa_cat.values()))
        existentes = set(Categoria.objects.values_list("nombre", flat=True))
        for c in cats:
            self.stdout.write(
                f"  - Categoría '{c}': {'existe' if c in existentes else 'se creará'}"
            )

    # -- escritura ------------------------------------------------------------

    def _escribir(self, mapa_cat):
        # Categorías
        categorias = {}
        for nombre in sorted(set(mapa_cat.values())):
            m = re.search(r"\d+", nombre)
            orden = int(m.group()) if m else None
            cat, _ = Categoria.objects.get_or_create(nombre=nombre, defaults={"orden": orden})
            categorias[nombre] = cat

        # Periodos (compartidos)
        periodos = {}
        for orden_p, nombre_p, *_ in PERIODOS:
            per, _ = Periodo.objects.get_or_create(
                nombre=nombre_p,
                vigencia=VIGENCIA,
                defaults={"orden": orden_p, "activo": False, "publico": True},
            )
            periodos[nombre_p] = per

        for m in self.modelos:
            modelo, _ = ModeloEvaluacion.objects.get_or_create(
                nombre=m["nombre"], version=VERSION, defaults={"activo": False}
            )
            _rep_dep, rep_pilares = m["representativa"]
            indice = self._construir_catalogo(modelo, rep_pilares)

            for dep_nombre, pilares in m["miembros"]:
                dependencia, _ = Dependencia.objects.get_or_create(nombre=dep_nombre)
                DependenciaModelo.objects.get_or_create(
                    modelo=modelo, dependencia=dependencia, defaults={"activo": False}
                )
                cat = categorias.get(mapa_cat.get(dep_nombre))
                periodos_dep = {
                    res["periodo"]
                    for p in pilares
                    for i in p["indicadores"]
                    for s in i["subindicadores"]
                    for res in s["resultados"]
                }
                evals = {}
                for nombre_p in periodos_dep:
                    ev, _ = Evaluacion.objects.get_or_create(
                        periodo=periodos[nombre_p],
                        dependencia=dependencia,
                        defaults={"modelo_evaluacion": modelo, "categoria": cat},
                    )
                    evals[nombre_p] = ev
                for p in pilares:
                    for i in p["indicadores"]:
                        for s in i["subindicadores"]:
                            sub_obj = indice[(p["nombre"], i["nombre"], s["nombre"])]
                            for res in s["resultados"]:
                                er, _ = EvaluacionResultado.objects.get_or_create(
                                    evaluacion=evals[res["periodo"]],
                                    subindicador=sub_obj,
                                    defaults={
                                        "puntaje": res["puntaje"],
                                        "ponderacion": res["ponderacion"],
                                        "observaciones": res["observaciones"],
                                    },
                                )
                                for d in res["detalles"]:
                                    EvaluacionResultadoDetalle.objects.get_or_create(
                                        resultado=er,
                                        mes=d["mes"],
                                        defaults={
                                            "puntaje": d["puntaje"],
                                            "ponderacion": d["ponderacion"],
                                        },
                                    )

    def _construir_catalogo(self, modelo, pilares):
        """Crea Pilar/Indicador/Subindicador/Criterio y devuelve índice por nombres."""
        indice = {}
        for p in pilares:
            pc, _ = PilarCategoria.objects.get_or_create(nombre=p["nombre"])
            pilar, _ = Pilar.objects.get_or_create(
                modelo_evaluacion=modelo,
                nombre=pc,
                defaults={"orden": p["orden"], "peso": p["peso"]},
            )
            for i in p["indicadores"]:
                ic, _ = IndicadorCategoria.objects.get_or_create(nombre=i["nombre"])
                indicador, _ = Indicador.objects.get_or_create(
                    pilar=pilar,
                    nombre=ic,
                    defaults={"orden": i["orden"], "peso": i["peso"]},
                )
                for s in i["subindicadores"]:
                    sc, _ = SubindicadorCategoria.objects.get_or_create(nombre=s["nombre"])
                    sub, _ = Subindicador.objects.get_or_create(
                        indicador=indicador,
                        nombre=sc,
                        defaults={
                            "orden": s["orden"],
                            "peso": s["peso"],
                            "tipo_calculo": s["tipo_calculo"],
                        },
                    )
                    indice[(p["nombre"], i["nombre"], s["nombre"])] = sub
                    for c in s["criterios"]:
                        Criterio.objects.get_or_create(
                            subindicador=sub,
                            nombre=c["nombre"],
                            defaults={"orden": c["orden"], "rango": c["rango"]},
                        )
        return indice
