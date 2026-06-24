# -*- coding: utf-8 -*-
"""
Agrega un PERIODO nuevo (p. ej. el trimestre Enero-Febrero-Marzo 2026) al modelo
**versión 1 ya existente**, sin recrear modelos: cada dependencia se vincula a su
`ModeloEvaluacion` v1 actual (vía `DependenciaModelo`) y se le agregan las evaluaciones
del periodo. Si una hoja trae un subindicador nuevo (p. ej. el de Estratégicos cambió
para 2026), se crea bajo su indicador en ese modelo.

Diferencias con `importar_v1` (la carga inicial): aquel agrupaba dependencias por
estructura y CREABA los modelos; este **reutiliza** los modelos existentes y solo añade
un periodo. El layout de columnas del periodo se **detecta desde los encabezados**
(`_detectar_periodos`), así sirve para bimestres/trimestres/etc.

Como en `importar_v1`: escala ×100, 5 decimales, la **ponderación se recalcula**
(`puntaje × peso`, no se usa la columna de ponderado del Excel) y se usa `update_or_create`.

Por defecto SIMULA (no escribe). Use --commit para aplicar.
"""
import re
from decimal import Decimal, ROUND_HALF_UP

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from contenido.models import (
    Categoria, Criterio, Dependencia, DependenciaModelo, Evaluacion,
    EvaluacionResultado, EvaluacionResultadoDetalle, Indicador, IndicadorCategoria,
    ModeloEvaluacion, Periodo, Pilar, PilarCategoria, Subindicador, SubindicadorCategoria,
)
from contenido.management.commands.importar_v1 import (
    limpiar_nombre, canon_pilar, a_porcentaje, norm_tipo_calculo, a_rango, _pond,
    Q5, CANON_DEPENDENCIA, VERSION, HOJA_CATEGORIAS,
)

MES_NUM = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}
MES_LABEL = {v: k.capitalize() for k, v in MES_NUM.items()}


def _h(ws, c):
    v = ws.cell(1, c).value
    return (limpiar_nombre(v) or "").lower() if v is not None else ""


def _detectar_periodos(ws):
    """Detecta los periodos desde los encabezados.

    Devuelve [(orden, nombre, [(mes, col_punt, col_pond)], col_pond_periodo|None, col_obs|None)].
    El nombre del periodo se arma con los meses que lo componen ("Enero - Febrero - Marzo").
    """
    maxc = ws.max_column
    periodos, cur, orden, c = [], [], 0, 9
    while c <= maxc:
        h = _h(ws, c)
        m = re.match(r"evaluacion_resultado_detalle_puntaje_mes_([a-záéíóú]+)$", h)
        if m and m.group(1) in MES_NUM:
            col_pond = c + 1 if _h(ws, c + 1).startswith("evaluacion_resultado_detalle_ponderado_mes_") else None
            cur.append((MES_NUM[m.group(1)], c, col_pond))
            c += 2 if col_pond else 1
            continue
        if h.startswith("ponderacion_"):
            col_obs = c + 1 if _h(ws, c + 1).startswith("evaluacion_resultado_observacion") else None
            if cur:
                orden += 1
                nombre = " - ".join(MES_LABEL[mn] for mn, _, _ in cur)
                periodos.append((orden, nombre, cur, c, col_obs))
                cur = []
            c += 2 if col_obs else 1
            continue
        if h.startswith("evaluacion_resultado_observacion"):
            # periodo de un solo mes sin columna de ponderación de periodo
            if cur:
                orden += 1
                nombre = " - ".join(MES_LABEL[mn] for mn, _, _ in cur)
                periodos.append((orden, nombre, cur, None, c))
                cur = []
            c += 1
            continue
        c += 1
    return periodos


def _limpiar_obs(v):
    return "" if v is None else str(v).replace("\xa0", " ").strip()


class Command(BaseCommand):
    help = "Agrega un periodo nuevo al modelo v1 existente. Simula salvo --commit."

    def add_arguments(self, parser):
        parser.add_argument("ruta")
        parser.add_argument("--vigencia", type=int, default=None,
                            help="Año del periodo (si se omite, se toma del nombre del archivo).")
        parser.add_argument("--commit", action="store_true")

    def handle(self, *args, **opts):
        import openpyxl
        ruta, commit = opts["ruta"], opts["commit"]
        try:
            wb = openpyxl.load_workbook(ruta, data_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f"No se encontró el archivo: {ruta}") from exc

        vigencia = opts["vigencia"]
        if vigencia is None:
            m = re.search(r"(20\d{2})", ruta)
            if not m:
                raise CommandError("No pude inferir la vigencia; pásala con --vigencia.")
            vigencia = int(m.group(1))

        modo = "ESCRITURA (--commit)" if commit else "SIMULACIÓN (dry-run)"
        self.stdout.write(self.style.MIGRATE_HEADING(f"\nImportar periodo (modelo v{VERSION}) — modo {modo}"))
        self.stdout.write(f"Archivo: {ruta}\nVigencia: {vigencia}")

        hojas = [ws for ws in wb.worksheets if ws.title.strip().lower() != HOJA_CATEGORIAS]
        # categorías (para la categoría de cada evaluación)
        mapa_cat = {}
        if HOJA_CATEGORIAS in [s.lower() for s in wb.sheetnames]:
            wsc = next(ws for ws in wb.worksheets if ws.title.strip().lower() == HOJA_CATEGORIAS)
            for r in range(2, wsc.max_row + 1):
                d = limpiar_nombre(wsc.cell(r, 1).value)
                cat = limpiar_nombre(wsc.cell(r, 2).value)
                if d:
                    mapa_cat[CANON_DEPENDENCIA.get(d, d)] = cat

        # periodo (de la primera hoja; todas comparten layout)
        periodos = _detectar_periodos(hojas[0])
        if not periodos:
            raise CommandError("No se detectó ningún periodo en los encabezados.")
        self.stdout.write(self.style.MIGRATE_LABEL(
            f"\nPeriodo(s) detectado(s): {[p[1] for p in periodos]}"))

        self._run(wb, hojas, mapa_cat, periodos, vigencia, commit)

    # -- núcleo --------------------------------------------------------------

    def _resolver_modelo(self, dep_nombre):
        dep = Dependencia.objects.filter(nombre=dep_nombre).first()
        if dep is None:
            return None, None
        dm = (DependenciaModelo.objects
              .filter(dependencia=dep, modelo__version=VERSION)
              .select_related("modelo").first())
        return dep, (dm.modelo if dm else None)

    def _run(self, wb, hojas, mapa_cat, periodos, vigencia, commit):
        nuevos_sub = []   # (modelo, pilar, ind, sub) que habría que crear
        nuevos_ind = []   # (modelo, pilar, ind) que habría que crear
        tot_eval = tot_er = tot_det = 0
        resumen = []

        @transaction.atomic
        def escribir():
            nonlocal tot_eval, tot_er, tot_det
            # categorías y periodos
            cat_obj = {}
            for nombre in sorted(set(v for v in mapa_cat.values() if v)):
                mm = re.search(r"\d+", nombre)
                cat_obj[nombre] = Categoria.objects.get_or_create(
                    nombre=nombre, defaults={"orden": int(mm.group()) if mm else None})[0]
            max_orden = (Periodo.objects.exclude(orden=None).order_by("-orden")
                         .values_list("orden", flat=True).first() or 0)
            per_obj = {}
            for i, (_o, nombre, *_rest) in enumerate(periodos):
                per_obj[nombre] = Periodo.objects.get_or_create(
                    nombre=nombre, vigencia=vigencia,
                    defaults={"orden": max_orden + 1 + i, "activo": False, "publico": True})[0]

            for ws in hojas:
                dep_nombre = CANON_DEPENDENCIA.get(ws.title.strip(), ws.title.strip())
                dep, modelo = self._resolver_modelo(dep_nombre)
                if modelo is None:
                    raise CommandError(f"«{dep_nombre}» no tiene modelo v{VERSION} en la BD.")
                cat = cat_obj.get(mapa_cat.get(dep_nombre))
                # índices del catálogo del modelo
                pil_idx = {}   # nombre -> Pilar
                ind_idx = {}   # (pilar,ind) -> Indicador
                sub_idx = {}   # (pilar,ind,sub) -> Subindicador
                for p in Pilar.objects.filter(modelo_evaluacion=modelo).select_related("nombre"):
                    pil_idx[p.nombre.nombre] = p
                for i in Indicador.objects.filter(pilar__modelo_evaluacion=modelo).select_related("nombre", "pilar__nombre"):
                    ind_idx[(i.pilar.nombre.nombre, i.nombre.nombre)] = i
                for s in Subindicador.objects.filter(indicador__pilar__modelo_evaluacion=modelo).select_related(
                        "nombre", "indicador__nombre", "indicador__pilar__nombre"):
                    sub_idx[(s.indicador.pilar.nombre.nombre, s.indicador.nombre.nombre, s.nombre.nombre)] = s

                for fila in self._parse(ws, periodos):
                    key = (fila["pilar"], fila["indicador"], fila["subindicador"])
                    sub = sub_idx.get(key)
                    if sub is None:   # subindicador (y quizá indicador/pilar) nuevo para este modelo
                        pilar = pil_idx.get(fila["pilar"])
                        if pilar is None:
                            pc, _ = PilarCategoria.objects.get_or_create(nombre=fila["pilar"])
                            pilar = Pilar.objects.create(
                                modelo_evaluacion=modelo, nombre=pc,
                                orden=fila["pilar_orden"], peso=fila["pilar_peso"] or Decimal("0"))
                            pil_idx[fila["pilar"]] = pilar
                        ind = ind_idx.get((fila["pilar"], fila["indicador"]))
                        if ind is None:
                            ic, _ = IndicadorCategoria.objects.get_or_create(nombre=fila["indicador"])
                            ind = Indicador.objects.create(
                                pilar=pilar, nombre=ic,
                                orden=fila["ind_orden"], peso=fila["ind_peso"] or Decimal("0"))
                            ind_idx[(fila["pilar"], fila["indicador"])] = ind
                        sc, _ = SubindicadorCategoria.objects.get_or_create(nombre=fila["subindicador"])
                        sub, _ = Subindicador.objects.get_or_create(
                            indicador=ind, nombre=sc,
                            defaults={"orden": fila["orden"], "peso": fila["peso"],
                                      "tipo_calculo": fila["tipo_calculo"]})
                        for c in fila["criterios"]:
                            Criterio.objects.get_or_create(
                                subindicador=sub, nombre=c["nombre"],
                                defaults={"orden": c["orden"], "rango": c["rango"]})
                        sub_idx[key] = sub
                    for res in fila["resultados"]:
                        ev, _ = Evaluacion.objects.get_or_create(
                            periodo=per_obj[res["periodo"]], dependencia=dep,
                            defaults={"modelo_evaluacion": modelo, "categoria": cat})
                        er, _ = EvaluacionResultado.objects.update_or_create(
                            evaluacion=ev, subindicador=sub,
                            defaults={"puntaje": res["puntaje"], "ponderacion": res["ponderacion"],
                                      "observaciones": res["observaciones"]})
                        for d in res["detalles"]:
                            EvaluacionResultadoDetalle.objects.update_or_create(
                                resultado=er, mes=d["mes"],
                                defaults={"puntaje": d["puntaje"], "ponderacion": d["ponderacion"]})

        # --- pasada de SIMULACIÓN (sin tocar BD): resumen y subs nuevos ---
        for ws in hojas:
            dep_nombre = CANON_DEPENDENCIA.get(ws.title.strip(), ws.title.strip())
            dep, modelo = self._resolver_modelo(dep_nombre)
            if modelo is None:
                resumen.append((dep_nombre, "SIN MODELO v1", 0, 0, 0)); continue
            existentes = set(
                Subindicador.objects.filter(indicador__pilar__modelo_evaluacion=modelo).values_list(
                    "indicador__pilar__nombre__nombre", "indicador__nombre__nombre", "nombre__nombre"))
            existentes_ind = set(
                Indicador.objects.filter(pilar__modelo_evaluacion=modelo).values_list(
                    "pilar__nombre__nombre", "nombre__nombre"))
            er = det = 0; pers = set()
            for fila in self._parse(ws, periodos):
                key = (fila["pilar"], fila["indicador"], fila["subindicador"])
                if (fila["pilar"], fila["indicador"]) not in existentes_ind:
                    nuevos_ind.append((modelo.nombre, fila["pilar"], fila["indicador"]))
                if key not in existentes:
                    nuevos_sub.append((modelo.nombre,) + key)
                for res in fila["resultados"]:
                    er += 1; det += len(res["detalles"]); pers.add(res["periodo"])
            tot_eval += len(pers); tot_er += er; tot_det += det
            resumen.append((dep_nombre, modelo.nombre, len(pers), er, det))

        self.stdout.write(self.style.MIGRATE_LABEL("\nDependencia -> modelo v1 (reusado):"))
        for dn, mn, np_, er, det in resumen:
            self.stdout.write(f"   {dn:24} -> {mn:34} {np_}p {er}res {det}det")
        if nuevos_ind:
            self.stdout.write(self.style.WARNING("\nIndicadores NUEVOS que se crearían (no estaban en v1):"))
            for x in sorted(set(nuevos_ind)):
                self.stdout.write(f"   + [{x[0]}] {x[1]} > {x[2]}")
        if nuevos_sub:
            self.stdout.write(self.style.WARNING("\nSubindicadores NUEVOS que se crearían (no estaban en v1):"))
            for x in sorted(set(nuevos_sub)):
                self.stdout.write(f"   + [{x[0]}] {x[1]} > {x[2]} > {x[3][:40]}")
        self.stdout.write(self.style.HTTP_INFO(
            f"\nTOTALES: {tot_eval} evaluaciones, {tot_er} resultados, {tot_det} detalles."))

        if not commit:
            self.stdout.write(self.style.WARNING("\n>>> SIMULACIÓN: no se escribió nada. Use --commit."))
            return
        escribir()
        self.stdout.write(self.style.SUCCESS("\n>>> Periodo importado."))

    def _parse(self, ws, periodos):
        """Genera dicts por subindicador (anclados en su fila) con pesos, criterios y resultados.

        Lleva también peso/orden de pilar e indicador (por si hay que crearlos en el modelo,
        p. ej. cuando un periodo agrega un indicador que no estaba en la v1)."""
        filas = []
        pil = ind = None
        pil_peso = pil_orden = ind_peso = ind_orden = None
        sub_actual = None
        orden_pilar = orden_ind = orden_sub = orden_crit = 0
        for r in range(2, ws.max_row + 1):
            pn = canon_pilar(limpiar_nombre(ws.cell(r, 1).value))
            inn = limpiar_nombre(ws.cell(r, 3).value)
            sn = limpiar_nombre(ws.cell(r, 5).value)
            cn = limpiar_nombre(ws.cell(r, 7).value)
            if pn:
                orden_pilar += 1
                pil, pil_peso, pil_orden = pn, a_porcentaje(ws.cell(r, 2).value), orden_pilar
                orden_ind = 0
            if inn:
                orden_ind += 1
                ind, ind_peso, ind_orden = inn, a_porcentaje(ws.cell(r, 4).value), orden_ind
                orden_sub = 0
            if sn:
                orden_sub += 1
                peso = a_porcentaje(ws.cell(r, 6).value)
                sub_actual = {
                    "pilar": pil, "indicador": ind, "subindicador": sn,
                    "pilar_peso": pil_peso, "pilar_orden": pil_orden,
                    "ind_peso": ind_peso, "ind_orden": ind_orden,
                    "orden": orden_sub, "peso": peso,
                    "tipo_calculo": norm_tipo_calculo(ws.cell(r, 17).value),
                    "criterios": [], "resultados": self._resultados(ws, r, peso, periodos),
                }
                filas.append(sub_actual)
                orden_crit = 0
            if cn and sub_actual is not None:
                orden_crit += 1
                sub_actual["criterios"].append(
                    {"orden": orden_crit, "nombre": cn, "rango": a_rango(ws.cell(r, 8).value)})
        return filas

    def _resultados(self, ws, fila, peso, periodos):
        out = []
        for orden_p, nombre_p, meses, col_pond, col_obs in periodos:
            detalles = []
            for mes, col_pun, col_pon in meses:
                pun = ws.cell(fila, col_pun).value
                pon = ws.cell(fila, col_pon).value if col_pon else None
                if pun is None and pon is None:
                    continue
                punt = a_porcentaje(pun if pun is not None else 0)
                detalles.append({"mes": mes, "puntaje": punt, "ponderacion": _pond(punt, peso)})
            pond_periodo = ws.cell(fila, col_pond).value if col_pond else None
            if not detalles and pond_periodo is None:
                continue
            if detalles:
                prom = sum((d["puntaje"] for d in detalles), Decimal("0")) / len(detalles)
                puntaje = prom.quantize(Q5, rounding=ROUND_HALF_UP)
            else:
                puntaje = Decimal("0.00000")
            out.append({
                "orden": orden_p, "periodo": nombre_p,
                "puntaje": puntaje, "ponderacion": _pond(puntaje, peso),
                "observaciones": _limpiar_obs(ws.cell(fila, col_obs).value) if col_obs else "",
                "detalles": detalles,
            })
        return out
