# -*- coding: utf-8 -*-
"""
Generación de informes de evaluación por dependencia (Excel y PDF).

Estos constructores reciben datos **planos** (los arma `construir_matriz` en
`views.py`) y no tocan el ORM ni las vistas: así se evita el import circular
(views importa este módulo de forma perezosa). Cada dependencia va en una hoja
(Excel) o página (PDF) distinta, con el logo de la Gobernación arriba a la
derecha y el encabezado institucional (Categoría · Dependencia · Periodo ·
Versión), replicando la jerarquía Pilar → Indicador → Subindicador → Criterios
de la pantalla de evaluación, con desglose por mes, puntaje y ponderación.

Motores: **openpyxl** (Excel) y **fpdf2** (PDF, Python puro). El logo se embebe
con manejo *best-effort*: si Pillow no está disponible (p. ej. bloqueado por una
política del SO), el informe se genera igual, solo sin la imagen.

Paleta y tipografía: tokens del Manual de Identidad (ver GUIA_DISEÑO.md).
"""
import re
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO

from django.contrib.staticfiles import finders

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- Paleta institucional (GUIA_DISEÑO.md) ---------------------------------
VERDE_SUP = "0e4d2a"
VERDE_PRIM = "109d39"
VERDE_CLARO = "e3f4ea"
AZUL = "0b72ab"
GRIS_CLARO = "dfe0e1"
GRIS_OSC = "5a595d"
BLANCO = "ffffff"

# equivalentes RGB para fpdf2
RGB_VERDE_SUP = (14, 77, 42)
RGB_VERDE_PRIM = (16, 157, 57)
RGB_PILAR = (205, 235, 214)
RGB_IND = (233, 244, 234)
RGB_AZUL = (11, 114, 171)
RGB_GRIS = (90, 89, 93)
RGB_GRIS_CLARO = (223, 224, 225)
# Separador entre pilares: mismo ancho (0.2) que el resto de la grilla, pero un gris algo más
# oscuro para que la línea SÍ se vea sobre el relleno de las columnas Pilar/Indicador (el gris
# claro de la grilla queda invisible ahí porque tiene casi la misma luminancia que el verde).
RGB_SEP_PILAR = (150, 152, 156)
SEP_PILAR_W = 0.2

# Logo institucional del encabezado (color, fondo claro). El lockup de la Secretaría de
# Planeación YA incluye el escudo y el texto "Gobernación de Sucre", así que se usa solo
# ese (no se repite el logo de Gobernación). La lista admite varios por si se agregan más.
LOGOS_REL = [
    "logos/logo-planeacion.png",  # Gobernación de Sucre · Secretaría de Planeación
]
MONT_REG = "fonts/Montserrat-Regular.ttf"
MONT_BOLD = "fonts/Montserrat-Bold.ttf"
NUMFMT = "0.00###"  # mínimo 2 decimales, hasta 5 (igual que el filtro `decimales`)

_MESES_ES = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
             "agosto", "septiembre", "octubre", "noviembre", "diciembre"]


def _logo_combinado(alto_px=220, sep_px=70):
    """Compone los logos institucionales en UNA imagen horizontal (lado a lado, misma
    altura, fondo transparente) para embeber idéntico en Excel y PDF. Devuelve una imagen
    PIL (RGBA) o None si Pillow no está disponible o no se encuentra ningún archivo
    (best-effort: el informe se genera igual sin logos)."""
    try:
        from PIL import Image
    except Exception:
        return None
    paths = [p for p in (finders.find(rel) for rel in LOGOS_REL) if p]
    if not paths:
        return None
    try:
        piezas = []
        for p in paths:
            im = Image.open(p).convert("RGBA")
            w = max(1, round(alto_px * im.width / im.height))
            piezas.append(im.resize((w, alto_px), Image.LANCZOS))
        ancho = sum(p.width for p in piezas) + sep_px * (len(piezas) - 1)
        lienzo = Image.new("RGBA", (ancho, alto_px), (0, 0, 0, 0))
        x = 0
        for pz in piezas:
            lienzo.alpha_composite(pz, (x, 0))
            x += pz.width + sep_px
        return lienzo
    except Exception:
        return None


def _fmt(value):
    """Número a texto con mínimo 2 y hasta 5 decimales, coma decimal (es-col)."""
    if value is None:
        return ""
    d = value if isinstance(value, Decimal) else Decimal(str(value))
    exp = d.normalize().as_tuple().exponent
    usados = -exp if isinstance(exp, int) and exp < 0 else 0
    places = max(2, min(5, usados))
    return ("{:.%df}" % places).format(d).replace(".", ",")


def _pct(value):
    """Peso a texto con exactamente 2 decimales, coma decimal (es-col), con %."""
    if value is None:
        return ""
    d = value if isinstance(value, Decimal) else Decimal(str(value))
    return "{:.2f}".format(d).replace(".", ",") + "%"


def _num(value):
    """Decimal/None a float para celdas numéricas de Excel (None -> None)."""
    return float(value) if value is not None else None


def _criterios_txt(criterios):
    return "\n".join("{}: {}".format(n, r) for n, r in criterios) if criterios else ""


def _total_ponderado(it):
    """Puntaje total de la dependencia = suma de la ponderación de todos los
    subindicadores (los sin resultado no suman). Devuelve un Decimal."""
    total = Decimal("0")
    for pilar in it["pilares"]:
        for ind in pilar["indicadores"]:
            for sub in ind["subindicadores"]:
                p = sub.get("ponderacion")
                if p is not None:
                    total += p if isinstance(p, Decimal) else Decimal(str(p))
    return total


# ===========================================================================
# EXCEL
# ===========================================================================
_XL_INVALIDOS = set('[]:*?/\\')


def _sheet_title(nombre, usados):
    """Título de hoja válido: sin caracteres prohibidos, ≤31 chars y único."""
    base = "".join(c for c in (nombre or "Hoja") if c not in _XL_INVALIDOS).strip() or "Hoja"
    base = base[:31]
    titulo, i = base, 2
    while titulo.lower() in usados:
        sufijo = " ({})".format(i)
        titulo = base[:31 - len(sufijo)] + sufijo
        i += 1
    usados.add(titulo.lower())
    return titulo


def generar_excel(items):
    """items: lista de dicts de `construir_matriz`. Devuelve un BytesIO (.xlsx)."""
    wb = Workbook()
    wb.remove(wb.active)
    # Logo combinado como PNG (bytes) para crear un XLImage nuevo por hoja.
    logo_png = None
    li = _logo_combinado()
    if li is not None:
        buf = BytesIO()
        li.save(buf, format="PNG")
        logo_png = buf.getvalue()
    usados = set()
    for it in items:
        ws = wb.create_sheet(_sheet_title(it["dependencia"], usados))
        _excel_hoja(ws, it, logo_png)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _head_cell(ws, r, c, texto, fill, alignment, borde):
    cell = ws.cell(r, c, texto)
    cell.font = Font(bold=True, color=BLANCO, size=10)
    cell.fill = fill
    cell.alignment = alignment
    cell.border = borde
    return cell


def _excel_hoja(ws, it, logo_png):
    meses = it["meses"]
    n_meses = len(meses)
    # Columnas (calca la evaluación): 1 Pilar | 2 Indicador | 3 Subindicador |
    # 4 Criterios | 5..(4+n) meses (bajo "Puntaje 0-100") | (5+n) Ponderación | Observaciones.
    mes0, mesN = 5, 4 + n_meses
    col_pond = 5 + n_meses
    col_obs = col_pond + 1
    n_cols = col_obs

    thin = Side(style="thin", color=GRIS_CLARO)
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_head = PatternFill("solid", fgColor=VERDE_SUP)
    fill_pilar = PatternFill("solid", fgColor=VERDE_CLARO)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- Encabezado institucional (filas 1-5) + logo arriba a la derecha ---
    ws["A1"] = "INFORME · MODELO DE ALTA GERENCIA"
    ws["A1"].font = Font(size=14, bold=True, color=VERDE_SUP)
    info = [
        ("Categoría", it["categoria"]),
        ("Dependencia", it["dependencia"]),
        ("Periodo", "{}{}".format(it["periodo"], " · {}".format(it["vigencia"]) if it["vigencia"] else "")),
        ("Versión", str(it["version"])),
    ]
    for i, (k, v) in enumerate(info, start=2):
        ws.cell(i, 1, "{}:".format(k)).font = Font(bold=True, color=GRIS_OSC, size=10)
        ws.cell(i, 2, v).font = Font(color=GRIS_OSC, size=10)
    if logo_png:
        try:
            img = XLImage(BytesIO(logo_png))  # BytesIO nuevo por hoja (no se comparte)
            ratio = (img.width / img.height) if img.height else 5
            img.height = 52
            img.width = int(52 * ratio)
            ws.add_image(img, "{}1".format(get_column_letter(mes0)))  # banda derecha del encabezado
        except Exception:
            pass  # sin Pillow: informe igual, sin logos

    # --- Cabecera de la tabla en 2 filas (7-8) ---
    hr, hr2 = 7, 8
    fijos = ["Pilar", "Indicador", "Subindicador", "Criterios"]
    for c, texto in enumerate(fijos, start=1):
        ws.merge_cells(start_row=hr, start_column=c, end_row=hr2, end_column=c)
        _head_cell(ws, hr, c, texto, fill_head, center, borde)
    if n_meses > 1:
        ws.merge_cells(start_row=hr, start_column=mes0, end_row=hr, end_column=mesN)
    _head_cell(ws, hr, mes0, "Puntaje (0-100)", fill_head, center, borde)
    for mi, (_mnum, lbl) in enumerate(meses):
        _head_cell(ws, hr2, mes0 + mi, lbl, fill_head, center, borde)
    ws.merge_cells(start_row=hr, start_column=col_pond, end_row=hr2, end_column=col_pond)
    _head_cell(ws, hr, col_pond, "Ponderación", fill_head, center, borde)
    ws.merge_cells(start_row=hr, start_column=col_obs, end_row=hr2, end_column=col_obs)
    _head_cell(ws, hr, col_obs, "Observaciones", fill_head, center, borde)
    for c in range(1, n_cols + 1):  # fill/borde en celdas cubiertas de la cabecera
        for rr in (hr, hr2):
            cell = ws.cell(rr, c)
            cell.fill = fill_head
            cell.border = borde

    # --- Filas de datos con celdas combinadas por Pilar/Indicador ---
    r = hr2 + 1
    for pilar in it["pilares"]:
        pilar_ini = r
        indicadores = pilar["indicadores"] or [None]
        for ind in indicadores:
            ind_ini = r
            subs = (ind["subindicadores"] if ind else []) or [None]
            for sub in subs:
                if sub:
                    nombre_sub = sub["nombre"]
                    if sub["peso"] is not None:
                        nombre_sub = "{}\n(peso {})".format(sub["nombre"], _pct(sub["peso"]))
                    ws.cell(r, 3, nombre_sub).alignment = left
                    ws.cell(r, 4, _criterios_txt(sub["criterios"])).alignment = left
                    if sub["tipo"] == "mensual":
                        for mi, (mnum, _lbl) in enumerate(meses):
                            cc = ws.cell(r, mes0 + mi, _num(sub["meses"].get(mnum)))
                            cc.number_format = NUMFMT; cc.alignment = center
                    else:  # directo: combina las celdas de los meses y muestra el puntaje
                        if n_meses > 1:
                            ws.merge_cells(start_row=r, start_column=mes0, end_row=r, end_column=mesN)
                        cc = ws.cell(r, mes0, _num(sub["puntaje"]))
                        cc.number_format = NUMFMT; cc.alignment = center
                    co = ws.cell(r, col_pond, _num(sub["ponderacion"]))
                    co.number_format = NUMFMT; co.alignment = center
                    ws.cell(r, col_obs, sub.get("observaciones", "")).alignment = left
                else:
                    ws.cell(r, 3, "—").alignment = left
                for c in range(1, n_cols + 1):
                    ws.cell(r, c).border = borde
                r += 1
            if ind:
                _merge_v(ws, 2, ind_ini, r - 1,
                         "{}\n(peso {})".format(ind["nombre"], _pct(ind["peso"])), center, borde)
        _merge_v(ws, 1, pilar_ini, r - 1,
                 "{}\n(peso {})".format(pilar["nombre"], _pct(pilar["peso"])), center, borde,
                 fill=fill_pilar)

    # --- Fila de total: puntaje de la dependencia = suma de la ponderación ---
    fill_total = PatternFill("solid", fgColor=VERDE_CLARO)  # verde bajo
    font_total = Font(bold=True, size=10, color=VERDE_SUP)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_pond - 1)
    lbl = ws.cell(r, 1, "PUNTAJE TOTAL DE LA DEPENDENCIA")
    lbl.font = font_total
    lbl.alignment = Alignment(horizontal="left", vertical="center")
    total = _total_ponderado(it).quantize(Decimal("0.01"), ROUND_HALF_UP)
    ct = ws.cell(r, col_pond, _num(total))
    ct.number_format = "0.00"  # puntaje total con exactamente 2 decimales
    ct.font = font_total
    ct.alignment = center
    for c in range(1, n_cols + 1):  # relleno + borde en toda la fila (incl. celdas combinadas)
        cell = ws.cell(r, c)
        cell.fill = fill_total
        cell.border = borde

    # --- Anchos de columna ---
    for col, w in (("A", 20), ("B", 22), ("C", 30), ("D", 34)):
        ws.column_dimensions[col].width = w
    for mi in range(n_meses):
        ws.column_dimensions[get_column_letter(mes0 + mi)].width = 11
    ws.column_dimensions[get_column_letter(col_pond)].width = 13
    ws.column_dimensions[get_column_letter(col_obs)].width = 46
    ws.freeze_panes = ws.cell(hr2 + 1, 1)
    ws.sheet_view.showGridLines = False


def _merge_v(ws, col, r0, r1, texto, alignment, borde, fill=None):
    if r1 < r0:
        return
    if r1 > r0:
        ws.merge_cells(start_row=r0, start_column=col, end_row=r1, end_column=col)
    cell = ws.cell(r0, col, texto)
    cell.alignment = alignment
    cell.font = Font(bold=True, size=10, color=(VERDE_SUP if fill is not None else GRIS_OSC))
    for rr in range(r0, r1 + 1):
        c = ws.cell(rr, col)
        c.border = borde
        if fill is not None:
            c.fill = fill


# ===========================================================================
# PDF  (fpdf2 — Python puro; logo best-effort)
# ===========================================================================
def _lat1(s):
    """Sanea texto al set latin-1 de las fuentes core de fpdf2 (evita errores por
    caracteres no soportados; devuelve '?' para los que falten)."""
    return str(s).encode("latin-1", "replace").decode("latin-1")


def _cap(texto, n=1400):
    """Tope de longitud por celda: evita que un texto enorme (una observación de
    varios párrafos) genere una fila más alta que la página (fpdf2 no la puede
    partir → error). n es holgado; el texto real no se ve afectado."""
    texto = texto or ""
    return texto if len(texto) <= n else texto[:n].rstrip() + "..."


# Emojis/pictogramas que ninguna fuente del PDF renderiza (saldrían como cuadro).
_EMOJI_RE = re.compile(
    "[\U0001F000-\U0001FAFF\U00002600-\U000027BF\U00002B00-\U00002BFF\U0000FE00-\U0000FE0F]"
)


def _t(pdf, s):
    """Texto para el PDF: quita emojis (no hay glifo) y, si la fuente incrustada soporta
    Unicode (Montserrat) lo usa tal cual; con la core de fallback (Helvetica) sanea latin-1."""
    s = _EMOJI_RE.sub("", str(s))
    return s if getattr(pdf, "unicode_ok", False) else _lat1(s)


def _fecha_generacion():
    """'Generado el 1 de julio de 2026, 14:35' (hora local del proyecto)."""
    try:
        from django.utils import timezone
        ahora = timezone.localtime()
    except Exception:
        ahora = datetime.now()
    return "Generado el {} de {} de {}, {:02d}:{:02d}".format(
        ahora.day, _MESES_ES[ahora.month - 1], ahora.year, ahora.hour, ahora.minute)


def _pdf_base():
    """FPDF con Montserrat incrustada (fallback a Helvetica) y pie de página con
    la fecha de generación + numeración."""
    from fpdf import FPDF

    class _InformePDF(FPDF):
        familia = "Helvetica"
        unicode_ok = False
        pie_fecha = ""

        def footer(self):
            self.set_y(-11)
            self.set_font(self.familia, "", 7)
            self.set_text_color(*RGB_GRIS)
            self.set_draw_color(*RGB_GRIS_CLARO)
            self.set_line_width(0.2)
            self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
            self.ln(1.5)
            half = (self.w - self.l_margin - self.r_margin) / 2
            self.cell(half, 5, self.pie_fecha, align="L")
            self.cell(half, 5, "Página {} de {{nb}}".format(self.page_no()), align="R")

    pdf = _InformePDF(orientation="L", unit="mm", format="A4")
    reg, bold = finders.find(MONT_REG), finders.find(MONT_BOLD)
    if reg and bold:
        try:
            pdf.add_font("Montserrat", "", reg)
            pdf.add_font("Montserrat", "B", bold)
            pdf.familia = "Montserrat"
            pdf.unicode_ok = True
        except Exception:
            pass  # fallback a Helvetica (core) + saneo latin-1
    return pdf


def generar_pdf(items):
    """items: lista de dicts de `construir_matriz`. Devuelve bytes (PDF)."""
    pdf = _pdf_base()
    pdf.pie_fecha = _t(pdf, _fecha_generacion())
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(True, margin=14)
    pdf.set_margins(12, 10, 12)
    logo = _logo_combinado()  # imagen PIL (o None) con ambos logos lado a lado
    for it in items:
        pdf.add_page()
        _pdf_encabezado(pdf, it, logo)
        _pdf_tabla(pdf, it)
    return bytes(pdf.output())


def _pdf_encabezado(pdf, it, logo):
    top = pdf.get_y()
    if logo is not None:
        try:
            h = 13.0
            w = h * (logo.width / logo.height)
            pdf.image(logo, x=pdf.w - pdf.r_margin - w, y=top, h=h)  # alineado a la derecha
        except Exception:
            pass  # sin Pillow: informe igual, sin logo
    pdf.set_xy(pdf.l_margin, top)
    pdf.set_font(pdf.familia, "B", 13)
    pdf.set_text_color(*RGB_VERDE_SUP)
    pdf.cell(0, 7, _t(pdf, "INFORME · MODELO DE ALTA GERENCIA"), new_x="LMARGIN", new_y="NEXT")

    per = "{}{}".format(it["periodo"], " · {}".format(it["vigencia"]) if it["vigencia"] else "")
    lineas = [
        ("Categoría", it["categoria"]), ("Dependencia", it["dependencia"]),
        ("Periodo", per), ("Versión", str(it["version"])),
    ]
    pdf.set_text_color(*RGB_GRIS)
    for k, v in lineas:
        pdf.set_font(pdf.familia, "B", 9)
        pdf.cell(pdf.get_string_width(_t(pdf, k + ": ")) + 1, 5, _t(pdf, k + ":"))
        pdf.set_font(pdf.familia, "", 9)
        pdf.cell(0, 5, " " + _t(pdf, v), new_x="LMARGIN", new_y="NEXT")

    y = pdf.get_y() + 1
    pdf.set_draw_color(*RGB_VERDE_PRIM)
    pdf.set_line_width(0.5)
    pdf.line(pdf.l_margin, y, pdf.w - pdf.r_margin, y)
    pdf.ln(4)


def _bordes_jerarquia(pil_starts, ind_starts, n_head, full_rows=frozenset(),
                      pil_sep=frozenset()):
    """Borde a medida para fpdf2: en las columnas Pilar (0) e Indicador (1) oculta las
    líneas horizontales *internas* de cada grupo, dejando solo la de inicio y fin del
    grupo. Combinado con el relleno de color continuo y el rótulo en la primera fila,
    esas celdas se ven **combinadas** (jerarquía clara) sin usar `rowspan` —que fpdf2 no
    puede partir entre páginas—. Las demás columnas conservan todos los bordes;
    `full_rows` fuerza todos los bordes en filas concretas (p. ej. la de total).
    `pil_sep` = filas que inician un nuevo pilar: su borde superior se dibuja como
    **separador de pilar** (todo el ancho) con un gris que sí contrasta con el relleno
    verde/azulado (el gris claro de la grilla ahí es invisible)."""
    from fpdf.enums import TableBordersLayout, TableBorderStyle, TableCellStyle

    sep = TableBorderStyle(color=RGB_SEP_PILAR, thickness=SEP_PILAR_W)

    class _Bordes(TableBordersLayout):
        def cell_style_getter(self, row_idx, col_idx, col_pos, num_heading_rows,
                              num_rows, num_col_idx, num_col_pos):
            top_sep = sep if row_idx in pil_sep else None  # separador entre pilares
            if row_idx < n_head or col_idx > 1 or row_idx in full_rows:
                return TableCellStyle(left=True, bottom=True, right=True,
                                      top=top_sep if top_sep is not None else True)
            starts = pil_starts if col_idx == 0 else ind_starts
            top = top_sep if top_sep is not None else (row_idx in starts)
            return TableCellStyle(
                left=True, right=True,
                top=top,                                                 # borde/separador al iniciar grupo
                bottom=(row_idx + 1 >= num_rows) or ((row_idx + 1) in starts),  # y al terminar
            )

    return _Bordes()


def _filas_rotulo(pilares):
    """Fila (índice base 0 dentro del cuerpo de la tabla) donde va el rótulo de cada
    Pilar/Indicador, centrado por **estructura** para que la jerarquía quede alineada:

      - **Indicador** → su subindicador **medio, sesgado hacia arriba**
        (`inicio + (n_subindicadores - 1) // 2`): con número par elige el **superior** de los
        dos centrales; con impar queda en el medio exacto.
      - **Pilar** → el rótulo de su indicador **medio, también sesgado hacia arriba**
        (`(n_indicadores - 1) // 2`), de modo que el rótulo del pilar quede a la altura del de
        su indicador central (par → el superior de los dos centrales).

    Replica el mismo aplanado que `_pdf_tabla` (un `None` cuando un pilar/indicador no tiene
    hijos, que igual ocupa una fila). Devuelve `(pil_label, ind_label)` como conjuntos de
    índices de fila."""
    pil_label, ind_label = set(), set()
    r = 0
    for pilar in pilares:
        inds = pilar["indicadores"] or [None]
        filas_ind = []  # fila del rótulo de cada indicador del pilar
        for ind in inds:
            subs = (ind["subindicadores"] if ind else []) or [None]
            fila = r + (len(subs) - 1) // 2  # medio sesgado hacia arriba (par → superior)
            ind_label.add(fila)
            filas_ind.append(fila)
            r += len(subs)
        pil_label.add(filas_ind[(len(filas_ind) - 1) // 2])  # indicador medio (sesgo arriba)
    return pil_label, ind_label


def _pdf_tabla(pdf, it):
    from fpdf.fonts import FontFace

    meses = it["meses"]
    n_meses = len(meses)
    fijos = ["Pilar", "Indicador", "Subindicador", "Criterios"]

    usable = pdf.w - pdf.l_margin - pdf.r_margin
    base = [24, 28, 34, 44] + [14] * n_meses + [16, 51]
    factor = usable / sum(base)
    widths = [w * factor for w in base]
    line_h = 3.8  # alto por línea de la tabla (debe coincidir con pdf.table(line_height=...))

    est_head = FontFace(emphasis="BOLD", color=(255, 255, 255), fill_color=RGB_VERDE_SUP)
    est_pilar = FontFace(emphasis="BOLD", color=RGB_VERDE_SUP, fill_color=RGB_PILAR)
    est_ind = FontFace(emphasis="BOLD", color=RGB_AZUL, fill_color=RGB_IND)
    est_total = FontFace(emphasis="BOLD", color=RGB_VERDE_SUP, fill_color=RGB_PILAR)  # verde bajo

    # Aplana la jerarquía en filas (una por subindicador) y guarda dónde empieza cada grupo
    # Pilar/Indicador para el borde a medida (fpdf2 no puede partir un rowspan entre páginas,
    # así que el rótulo va en una sola fila y el "merge" se simula con bordes+relleno).
    filas = []
    pil_starts, ind_starts = set(), set()
    N_HEAD = 2  # la cabecera ocupa 2 filas
    for pilar in it["pilares"]:
        inds = pilar["indicadores"] or [None]
        pil_starts.add(N_HEAD + len(filas))
        for ind in inds:
            ind_starts.add(N_HEAD + len(filas))
            for sub in (ind["subindicadores"] if ind else []) or [None]:
                filas.append((pilar, ind, sub))
    total_row_idx = N_HEAD + len(filas)  # la fila de total va al final, con todos los bordes
    pil_sep = pil_starts - {N_HEAD}  # separador entre pilares (todos los inicios menos el primero)
    bordes = _bordes_jerarquia(pil_starts, ind_starts, N_HEAD, {total_row_idx}, pil_sep)

    # Rótulo de Pilar/Indicador centrado por **estructura** (no por altura de píxeles): el
    # indicador en su subindicador medio y el pilar en el rótulo de su indicador medio, de modo
    # que el rótulo del pilar quede alineado con el de su indicador central.
    pil_label, ind_label = _filas_rotulo(it["pilares"])

    pdf.set_font(pdf.familia, "", 6.5)
    pdf.set_text_color(*RGB_GRIS)
    pdf.set_draw_color(*RGB_GRIS_CLARO)
    pdf.set_line_width(0.2)

    with pdf.table(col_widths=widths, line_height=line_h, first_row_as_headings=False,
                   text_align="CENTER", v_align="MIDDLE", borders_layout=bordes) as table:
        # Cabecera en 2 filas: los meses van bajo "Puntaje (0-100)".
        h1 = table.row()
        for t in fijos:
            h1.cell(_t(pdf, t), rowspan=2, style=est_head, align="CENTER")
        h1.cell(_t(pdf, "Puntaje (0-100)"), colspan=n_meses, style=est_head, align="CENTER")
        h1.cell(_t(pdf, "Ponderación"), rowspan=2, style=est_head, align="CENTER")
        h1.cell(_t(pdf, "Observaciones"), rowspan=2, style=est_head, align="CENTER")
        h2 = table.row()
        for _mnum, lbl in meses:
            h2.cell(_t(pdf, lbl), style=est_head, align="CENTER")

        # Pilar/Indicador: se rotula solo la fila elegida por `_filas_rotulo` (indicador en su
        # subindicador medio, pilar en su indicador medio) con v_align MIDDLE; el relleno de
        # color y el borde a medida hacen que la columna se vea como una celda combinada.
        for i, (pilar, ind, sub) in enumerate(filas):
            pilar_txt = _t(pdf, "{}\n(peso {})".format(pilar["nombre"], _pct(pilar["peso"])))
            ind_txt = _t(pdf, "{}\n(peso {})".format(ind["nombre"], _pct(ind["peso"]))) if ind else _t(pdf, "—")
            row = table.row()
            row.cell(pilar_txt if i in pil_label else "", style=est_pilar, v_align="MIDDLE")
            row.cell(ind_txt if i in ind_label else "", style=est_ind, v_align="MIDDLE")
            if sub:
                row.cell(_t(pdf, "{}\n(peso {})".format(sub["nombre"], _pct(sub["peso"]))))
                row.cell(_t(pdf, _cap(_criterios_txt(sub["criterios"]))))
                if sub["tipo"] == "mensual":
                    for mnum, _lbl in meses:
                        row.cell(_fmt(sub["meses"].get(mnum)), align="CENTER")
                else:  # directo: combina las celdas de los meses y muestra el puntaje
                    row.cell(_fmt(sub["puntaje"]), colspan=n_meses, align="CENTER")
                row.cell(_fmt(sub["ponderacion"]), align="CENTER")
                row.cell(_t(pdf, _cap(sub.get("observaciones", ""))))
            else:
                row.cell(_t(pdf, "—"))
                row.cell("")
                row.cell("", colspan=n_meses)
                row.cell("")
                row.cell("")

        # Fila de total: puntaje de la dependencia = suma de la ponderación (2 decimales)
        total = _total_ponderado(it).quantize(Decimal("0.01"), ROUND_HALF_UP)
        tr = table.row()
        tr.cell(_t(pdf, "PUNTAJE TOTAL DE LA DEPENDENCIA"), colspan=4 + n_meses,
                style=est_total, align="LEFT")
        tr.cell(_fmt(total), style=est_total, align="CENTER")
        tr.cell("", style=est_total)
