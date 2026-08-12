"""Fidelidad del informe: los puntajes/ponderaciones y la fila de total del
reporte deben coincidir con lo guardado en la base de datos.

Flujo del test:
  1. Se diligencia una evaluacion real (un subindicador directo + uno mensual),
     lo que crea los `EvaluacionResultado` con su `ponderacion` en la BD.
  2. Se genera el informe Excel via la vista `reporte_generar`.
  3. Se lee el .xlsx y se compara contra la BD:
       - cada `ponderacion` de la BD aparece como celda del reporte;
       - la fila "PUNTAJE TOTAL DE LA DEPENDENCIA" = suma de las ponderaciones.
  4. Ademas se verifica que Excel y PDF siguen generandose (el cambio no rompe).
"""
from decimal import Decimal
from io import BytesIO

from django.test import SimpleTestCase
from django.urls import reverse
from openpyxl import load_workbook

from contenido.models import EvaluacionResultado
from contenido.reportes import _filas_rotulo

from .base import BaseMagTestCase

TOTAL_LBL = "PUNTAJE TOTAL DE LA DEPENDENCIA"


def _pilar(*indicadores):
    """Helper: arma un dict de pilar con sus indicadores; cada indicador se pasa como
    el nº de subindicadores que tiene."""
    return {"nombre": "P", "peso": None, "indicadores": [
        {"nombre": "I", "peso": None,
         "subindicadores": [{"nombre": "S"} for _ in range(n)]}
        for n in indicadores
    ]}


class FilasRotuloTests(SimpleTestCase):
    """El rótulo de Pilar/Indicador se centra por ESTRUCTURA con el medio **sesgado hacia
    arriba**: indicador en `(n_subs - 1) // 2` y pilar en su indicador `(n_indicadores - 1) // 2`
    (par -> el superior de los dos centrales; impar -> medio exacto)."""

    def test_indicador_impar_cae_en_el_medio_exacto(self):
        # 5 subindicadores (impar) -> medio exacto = fila 2
        _pil, ind = _filas_rotulo([_pilar(5)])
        self.assertEqual(ind, {2})

    def test_indicador_par_cae_en_el_superior_de_los_dos_centrales(self):
        # 2 subs -> fila 0 (arriba, no la de abajo); 4 subs -> fila 1
        self.assertEqual(_filas_rotulo([_pilar(2)])[1], {0})
        self.assertEqual(_filas_rotulo([_pilar(4)])[1], {1})

    def test_pilar_se_alinea_con_su_indicador_medio(self):
        # 3 indicadores de 1 sub c/u -> filas ind 0,1,2; el pilar toma el del medio (fila 1)
        pil, ind = _filas_rotulo([_pilar(1, 1, 1)])
        self.assertEqual(ind, {0, 1, 2})
        self.assertEqual(pil, {1})

    def test_pilar_par_toma_el_indicador_superior_de_los_dos_centrales(self):
        # 2 indicadores de 1 sub c/u -> filas ind 0,1; el pilar toma el superior (fila 0)
        pil, ind = _filas_rotulo([_pilar(1, 1)])
        self.assertEqual(ind, {0, 1})
        self.assertEqual(pil, {0})

    def test_pilar_alineado_con_indicador_central_de_tamano_distinto(self):
        # indicadores de 2, 3 y 1 subs -> filas ind: 0+0=0, 2+1=3, 5+0=5; medio = fila 3
        pil, ind = _filas_rotulo([_pilar(2, 3, 1)])
        self.assertEqual(ind, {0, 3, 5})
        self.assertEqual(pil, {3})  # el rótulo del pilar coincide con el de su indicador medio

    def test_indices_son_globales_entre_pilares(self):
        # dos pilares de un indicador (1 sub) cada uno -> filas 0 y 1
        pil, ind = _filas_rotulo([_pilar(1), _pilar(1)])
        self.assertEqual(ind, {0, 1})
        self.assertEqual(pil, {0, 1})


class ReporteTotalesTests(BaseMagTestCase):
    def setUp(self):
        super().setUp()
        self.login_admin()
        self.evaluacion = self.crear_evaluacion()
        url = reverse("contenido:evaluacion_diligenciar", args=[self.evaluacion.pk])
        # Puntajes distintos para que las ponderaciones tambien difieran:
        #   directo 90  -> ponderacion 90 * 10 / 100 = 9
        #   mensual avg(60, 80) = 70 -> ponderacion 70 * 10 / 100 = 7
        self.client.post(url, {
            "puntaje_{}".format(self.sub_directo.pk): "90",
            "puntaje_{}_{}".format(self.sub_mensual.pk, 1): "60",
            "puntaje_{}_{}".format(self.sub_mensual.pk, 2): "80",
        })

    # ------------------------------------------------------------------ helpers
    def _generar(self, formato):
        resp = self.client.post(reverse("contenido:reporte_generar"), {
            "modelo": self.modelo.version, "categoria": self.categoria.pk,
            "periodo": self.periodo.pk, "formato": formato,
        })
        self.assertEqual(resp.status_code, 200, resp.content[:200])
        return resp

    def _ponderaciones_bd(self):
        return {
            r.subindicador_id: r.ponderacion
            for r in EvaluacionResultado.objects.filter(evaluacion=self.evaluacion)
        }

    # -------------------------------------------------------------------- tests
    def test_bd_tiene_las_ponderaciones_esperadas(self):
        """Sanity: la BD guardo lo que esperamos antes de comparar el reporte."""
        pond = self._ponderaciones_bd()
        self.assertEqual(pond[self.sub_directo.pk], Decimal("9.00000"))
        self.assertEqual(pond[self.sub_mensual.pk], Decimal("7.00000"))

    def test_reporte_coincide_con_bd_y_total_es_la_suma(self):
        resp = self._generar("excel")
        ws = load_workbook(BytesIO(resp.content)).worksheets[0]
        rows = [[c.value for c in row] for row in ws.iter_rows()]

        pond_bd = self._ponderaciones_bd()
        total_bd = sum(pond_bd.values(), Decimal("0"))

        # 1) cada ponderacion de la BD aparece como celda numerica del reporte
        numeros = {round(float(v), 5)
                   for r in rows for v in r if isinstance(v, (int, float))}
        for sid, pond in pond_bd.items():
            self.assertIn(round(float(pond), 5), numeros,
                          "falta la ponderacion del subindicador {}".format(sid))

        # 2) la fila de total lleva el rotulo y un unico valor = suma de la BD
        fila_total = next(
            (r for r in rows if TOTAL_LBL in [str(v) for v in r]), None
        )
        self.assertIsNotNone(fila_total, "no se genero la fila de total")
        valores = [v for v in fila_total if isinstance(v, (int, float))]
        self.assertEqual(len(valores), 1, "la fila de total debe tener un solo numero")
        self.assertAlmostEqual(valores[0], float(total_bd), places=5)
        # el total es realmente la suma (9 + 7 = 16), no un valor suelto
        self.assertAlmostEqual(valores[0], 16.0, places=5)

    def test_generacion_no_se_rompe_excel_y_pdf(self):
        """El cambio de la fila de total no rompe ninguno de los dos formatos."""
        xls = self._generar("excel")
        self.assertIn("spreadsheetml", xls["Content-Type"])
        self.assertTrue(load_workbook(BytesIO(xls.content)).worksheets)

        pdf = self._generar("pdf")
        self.assertEqual(pdf["Content-Type"], "application/pdf")
        self.assertTrue(pdf.content.startswith(b"%PDF"))
