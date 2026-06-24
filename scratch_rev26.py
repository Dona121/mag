# -*- coding: utf-8 -*-
import sys; sys.stdout.reconfigure(encoding="utf-8")
import openpyxl, re
from collections import defaultdict
from contenido.models import DependenciaModelo, Subindicador, Criterio, Dependencia

P = r"C:\Users\Donal\Documents\MEGA\Estudio\Django\modelo_alta_gerencia\migracion\estructura_modelo_version_1_2026.xlsx"
wb = openpyxl.load_workbook(P, data_only=True)
CANON={"Ciclo de Proyectos":"Ciclos de Gerencia","Mecanismos de Financiación":"Otros Mecanismos de Financiación"}
def cp(x): return CANON.get(x,x)
def clean(v): return re.sub(r"\s+"," ",str(v).replace("\xa0"," ")).strip() if v is not None else None
def fix_typo(s):
    if s is None: return s
    s=s.replace("cumplimeinto","cumplimiento"); s=re.sub(r"\bCumplimento\b","Cumplimiento",s); return s
def norm_name(v): return fix_typo(clean(v))
def f(v):
    try: return float(v)
    except (TypeError,ValueError): return None

# ---- DB v1: dependencia->modelo, y catálogo por modelo ----
dep_model = {dm.dependencia.nombre: dm.modelo_id
             for dm in DependenciaModelo.objects.filter(modelo__version=1).select_related("dependencia")}
db_cat = defaultdict(set)      # model_id -> set((pil,ind,sub))
db_crit = defaultdict(set)     # (model_id,pil,ind,sub) -> set(criterio)
sub_of = {}                    # sub_id -> (model_id,pil,ind,sub)
for r in Subindicador.objects.filter(indicador__pilar__modelo_evaluacion__version=1).values(
        "id","indicador__pilar__modelo_evaluacion_id",
        "indicador__pilar__nombre__nombre","indicador__nombre__nombre","nombre__nombre"):
    key=(r["indicador__pilar__modelo_evaluacion_id"], r["indicador__pilar__nombre__nombre"],
         r["indicador__nombre__nombre"], r["nombre__nombre"])
    db_cat[key[0]].add(key[1:]); sub_of[r["id"]]=key
for r in Criterio.objects.filter(subindicador__indicador__pilar__modelo_evaluacion__version=1).values(
        "subindicador_id","nombre"):
    k=sub_of.get(r["subindicador_id"])
    if k: db_crit[k].add(clean(r["nombre"]))

db_deps = set(Dependencia.objects.values_list("nombre", flat=True))

# ---- Parsear 2026 ----
def parse(ws):
    pil=ind=None; cat=defaultdict(set); crit=defaultdict(set); rows=[]
    for r in range(2, ws.max_row+1):
        pn=cp(norm_name(ws.cell(r,1).value)); inn=norm_name(ws.cell(r,3).value); sn=norm_name(ws.cell(r,5).value); cn=norm_name(ws.cell(r,7).value)
        if pn: pil=pn
        if inn: ind=inn
        if sn: cur=(pil,ind,sn); cat[ws.title].add(cur)
        if cn and sn is None and pil and ind: pass
        if cn: crit[(pil,ind,sn or cur[2])].add(cn)
        rows.append((r,pn,inn,sn,cn))
    return cat[ws.title], crit

print("=== [A] Dependencias 2026 vs BD ===")
hojas=[w for w in wb.worksheets if w.title.strip().lower()!="categorias"]
faltan=[]; 
for ws in hojas:
    nom=clean(ws.title)
    estado = "OK" if nom in db_deps else "NO EXISTE en BD"
    if nom not in db_deps: faltan.append(nom)
    print(f"   {nom:28} {estado}")
# categorias sheet
wscat=wb["categorias"]
print("\n   Hoja 'categorias':")
for r in range(2, wscat.max_row+1):
    d=clean(wscat.cell(r,1).value); c=clean(wscat.cell(r,2).value)
    if d: print(f"      {d:28} -> {c}   {'(dep no en BD)' if d not in db_deps else ''}")

print("\n=== [B] Catálogo 2026 vs catálogo v1 en BD (por dependencia) ===")
for ws in hojas:
    nom=clean(ws.title)
    mid=dep_model.get(nom)
    cat,crit = parse(ws)
    if mid is None:
        print(f"   {nom}: sin modelo v1 en BD (no comparable)"); continue
    extra = cat - db_cat[mid]      # en 2026, no en BD
    falta = db_cat[mid] - cat      # en BD, no en 2026
    cdiff=[]
    for k in (cat & db_cat[mid]):
        kk=(mid,)+k
        if crit.get(k,set()) != db_crit.get(kk,set()):
            cdiff.append(k[2][:30])
    msg = "OK (nombres y criterios coinciden con v1)" if not extra and not falta and not cdiff else ""
    print(f"   {nom}: {msg}")
    for x in sorted(extra): print(f"       + EXTRA en 2026: {x}")
    for x in sorted(falta): print(f"       - FALTA (está en v1, no en 2026): {x}")
    for x in cdiff: print(f"       ~ criterios distintos en: {x}")
